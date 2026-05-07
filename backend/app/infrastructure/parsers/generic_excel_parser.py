from io import BytesIO

import openpyxl

from app.domain.models.raw_transaction import RawTransaction


class GenericExcelParser:
    """
    Parses a standard Excel file where row 1 contains column headers:
    date, description, amount, type[, category]  (case-insensitive)
    Same signed-amount convention as CSVParser: expense → negative amount.
    """

    def parse(self, content: bytes, filename: str) -> list[RawTransaction]:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        ws = wb.active

        headers = [
            str(cell.value).strip().lower() if cell.value is not None else ""
            for cell in ws[1]
        ]
        col = {name: idx for idx, name in enumerate(headers)}

        if "date" not in col or "amount" not in col:
            return []

        results: list[RawTransaction] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                date_val = str(row[col["date"]] or "").strip()
                description = str(row[col.get("description", -1)] or "").strip() if "description" in col else ""
                raw_amount = float(row[col["amount"]])  # type: ignore[arg-type]
                tx_type = str(row[col.get("type", -1)] or "").strip().lower() if "type" in col else ""

                if tx_type == "expense" and raw_amount > 0:
                    raw_amount = -raw_amount

                if not date_val:
                    continue

                results.append(
                    RawTransaction(
                        external_id=None,
                        date_str=date_val,
                        sender="",
                        description=description,
                        amount=raw_amount,
                        source="generic_excel",
                    )
                )
            except Exception:
                continue

        return results
